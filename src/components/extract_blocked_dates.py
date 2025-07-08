import sys
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import re
from datetime import datetime
import unicodedata
import requests
import os

# 📥 Get Excel file path from command-line args
if len(sys.argv) < 2:
    print("❌ No Excel file path provided.")
    sys.exit(1)

EXCEL_PATH = sys.argv[1]

if not os.path.exists(EXCEL_PATH):
    print(f"❌ File not found: {EXCEL_PATH}")
    sys.exit(1)

# 📊 Load workbook
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# 📅 Collect all real date cells
date_cells = {}
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, datetime):
            date_cells[cell.coordinate] = cell.value
        elif isinstance(cell.value, (int, float)):
            try:
                dt = from_excel(cell.value, wb.epoch)
                date_cells[cell.coordinate] = dt
            except Exception:
                pass

calendar_dates = list(date_cells.values())

month_map = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Sept': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

linked_results = []

# 🔍 Match date ranges
def find_matching_dates_in_range(start_day, end_day, month, year_guess, remark_text, remark_cell):
    month_num = month_map[month]
    for dt in calendar_dates:
        if dt.year == year_guess and dt.month == month_num and start_day <= dt.day <= end_day:
            linked_results.append({
                "date": dt.strftime("%Y-%m-%d"),
                "remark": remark_text,
                "remark_cell": remark_cell
            })

# 📝 Parse remark column (H)
for row in ws.iter_rows():
    cell = row[7]  # Column H
    if cell.value and isinstance(cell.value, str):
        text = unicodedata.normalize("NFKC", cell.value.strip())
        text = re.sub(r'[\u2013\u2014\u2212]', '-', text)  # Normalize dashes
        cleaned_text = re.sub(
            r'^\s*\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sept|Sep|Oct|Nov|Dec)'
            r'(?:\s+\d{2,4})?'  # optional year
            r'(?:\s*(Mon|Tue|Wed|Thu|Fri|Sat|Sun))?'  # optional weekday
            r'[\s,:-]*',
            '',
            text,
            flags=re.UNICODE
        )

        # Case: Normal remarks
        if 'HOR Week' not in text and 'CBL Week' not in text:
            # Match single dates
            single_dates = re.findall(r'\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\b', text)
            for day, month in single_dates:
                for dt in calendar_dates:
                    if dt.day == int(day) and dt.month == month_map[month]:
                        linked_results.append({
                            "date": dt.strftime("%Y-%m-%d"),
                            "remark": cleaned_text,
                            "remark_cell": cell.coordinate
                        })

            # Match ranges like 12 - 15 Mar
            date_ranges = re.findall(r'(\d{1,2})\s*-\s*(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)', text)
            for start_day, end_day, month in date_ranges:
                year_guess = max(dt.year for dt in calendar_dates if dt.month == month_map[month])
                find_matching_dates_in_range(int(start_day), int(end_day), month, year_guess, cleaned_text, cell.coordinate)

            # Match full ranges like 12 Mar to 15 Mar
            full_range = re.findall(
                r'(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\s+(?:to|-)\s+(\d{1,2})\s+\2(?:\s+(\d{4}))?',
                text)
            for start_day, month, end_day, year in full_range:
                year = int(year) if year else datetime.now().year
                find_matching_dates_in_range(int(start_day), int(end_day), month, year, cleaned_text, cell.coordinate)

        # Case: HOR / CBL Weeks
        if 'HOR Week' in text or 'CBL Week' in text:
            for date_cell in row[:7]:
                if isinstance(date_cell.value, datetime):
                    linked_results.append({
                        "date": date_cell.value.strftime("%Y-%m-%d"),
                        "remark": text,
                        "remark_cell": cell.coordinate
                    })

# 🧼 Deduplicate
seen = set()
deduped_results = []
for item in linked_results:
    key = (item['date'], item['remark_cell'])
    if key not in seen:
        seen.add(key)
        deduped_results.append(item)

# 📦 Final payload
final_payload = [{"date": item["date"], "remark": item["remark"]} for item in deduped_results]

# 🚀 POST to backend
url = "http://localhost:3001/api/scheduling/update-blocked-dates"
headers = {"Content-Type": "application/json"}

print(f"📡 Sending {len(final_payload)} blocked dates to backend...")
try:
    response = requests.post(url, json={"blocked_dates": final_payload}, headers=headers)
    if response.status_code == 200:
        print("✅ Blocked dates successfully updated.")
    else:
        print("❌ Backend error:", response.status_code, response.text)
        sys.exit(1)
except Exception as e:
    print("❌ Failed to send data to backend:", e)
    sys.exit(1)
