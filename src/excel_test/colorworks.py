from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook("Copy of AY 2526 Phase III Medicine Posting Dates in Summary - For HCI Dated 21 Feb 25.xlsx")
ws = wb.active

colored_cells = []

for row in ws.iter_rows():
    for cell in row:
        fill = cell.fill
        if fill and fill.patternType == 'solid':
            color = fill.start_color

            # Check for RGB color that isn't white or transparent
            if color.type == 'rgb':
                rgb = color.rgb
                if rgb and rgb.upper() not in ['FFFFFFFF', '00000000', 'FF000000']:
                    colored_cells.append((cell.coordinate, cell.value, f"#{rgb[-6:]}"))

            # Check for theme-based colors that aren't white (theme 0)
            elif color.type == 'theme':
                if color.theme != 0:
                    colored_cells.append((cell.coordinate, cell.value, f"theme:{color.theme}"))

            # Check for indexed colors that aren't white or transparent
            elif color.type == 'indexed':
                if color.indexed not in [64, 9]:
                    colored_cells.append((cell.coordinate, cell.value, f"indexed:{color.indexed}"))

# Output final list of visibly colored cells
for coord, val, col in colored_cells:
    print(f"🟩 Cell {coord}: Value = {val}, Fill Color = {col}")
